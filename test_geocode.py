#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
片区编码精度评估脚本
从Excel读取地址，调用高德API编码并分类到片区，与已知标签对比。
运行: python3 test_geocode.py
"""
import sys, os, json, datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
os.chdir(os.path.dirname(os.path.abspath(__file__)))

from app import geocode_address, load_zones, point_in_polygon

# Excel路径
EXCEL_PATH = os.path.expanduser(r"D:\桌面\上城区单位信息汇总.xlsx")
# 如果UOS路径不同，手动修改上面这行

def main():
    out_lines = []
    def log(s=""):
        print(s)
        out_lines.append(s)

    log("=" * 70)
    log("  片区编码精度评估报告")
    log("  时间: " + datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    log("=" * 70)

    # 加载片区
    zones = load_zones()
    if not zones:
        log("❌ 未找到片区数据 (zones.json)")
        return
    log("  片区: %d 个" % len(zones))
    for z in zones:
        log("    - %s (%d 个顶点)" % (z["name"], len(z["polygon"])))

    # 读取Excel
    import openpyxl
    if not os.path.exists(EXCEL_PATH):
        log("\n❌ Excel文件不存在: %s" % EXCEL_PATH)
        log("   请修改 test_geocode.py 中的 EXCEL_PATH 变量。")
        # 退回到测试地址
        log("\n   改用内置测试地址。")
        test_addresses = [
            ("杭州市上城区唯康老人养生文化公寓", "杭州市上城区近江家园四园18幢", ""),
            ("杭州市崇文实验学校", "上城区近江南路1号", ""),
            ("杭州文化广播电视集团", "浙江省杭州市上城区之江路888号", ""),
            ("杭州祈嘉食品有限公司", "浙江省杭州市上城区景江城市花园2幢219室", ""),
            ("", "杭州市上城区清江路188号", ""),
            ("杭州娃哈哈集团有限公司", "浙江省杭州市上城区清泰街160号", ""),
        ]
    else:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        ws = wb.active
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        log("\n  Excel: %s" % os.path.basename(EXCEL_PATH))
        log("  表头: %s" % " | ".join(str(h) for h in headers))
        log("  行数: %d (含表头)" % ws.max_row)

        name_col = 1  # 单位名称
        addr_col = 2  # 地址
        zone_col = 3  # 片区

        test_addresses = []
        for r in range(2, ws.max_row + 1):
            name = ws.cell(r, name_col).value or ""
            addr = ws.cell(r, addr_col).value or ""
            expected_zone = ws.cell(r, zone_col).value or ""
            if addr:
                test_addresses.append((str(name), str(addr), str(expected_zone)))

    log("\n共 %d 个待测试地址\n" % len(test_addresses))
    log("-" * 70)

    results = []
    success = 0
    fail = 0

    for i, (name, addr, expected_zone) in enumerate(test_addresses, 1):
        log("[%d/%d] %s" % (i, len(test_addresses), (name[:30] + "..") if len(name) > 30 else name))
        log("      地址: %s" % addr[:50])

        result = geocode_address(addr, name)

        if result and result.get("lng"):
            lng, lat = result["lng"], result["lat"]
            strategy = result.get("strategy", "??")
            level = result.get("level", "")

            # 判断落在哪个片区
            matched_zone = "未划分"
            matched_color = None
            for z in zones:
                if point_in_polygon(lng, lat, z["polygon"]):
                    matched_zone = z["name"]
                    matched_color = z["color"]
                    break

            is_correct = matched_zone == expected_zone
            if is_correct:
                success += 1
                log("  ✅ (%.6f, %.6f) → %s   策略: %s" % (lng, lat, matched_zone, strategy))
            else:
                fail += 1
                tag = "❌" if expected_zone else "⚠️"
                log("  %s (%.6f, %.6f) → %s  期望: %s  策略: %s" % (
                    tag, lng, lat, matched_zone, expected_zone or "(未知)", strategy))

            results.append({
                "name": name, "addr": addr,
                "lng": lng, "lat": lat,
                "strategy": strategy, "level": level,
                "matched_zone": matched_zone,
                "expected_zone": expected_zone,
                "correct": is_correct,
            })
        else:
            fail += 1
            log("  ❌ 编码失败")
            results.append({
                "name": name, "addr": addr,
                "lng": None, "lat": None,
                "strategy": "FAILED", "level": "",
                "matched_zone": "N/A",
                "expected_zone": expected_zone,
                "correct": False,
            })

    # 汇总
    total = len(test_addresses)
    correct_labeled = sum(1 for r in results if r["expected_zone"])
    correct_matched = sum(1 for r in results if r["correct"])

    log("\n" + "=" * 70)
    log("  评估汇总")
    log("  测试总数: %d" % total)
    log("  编码成功: %d" % success)
    log("  编码失败: %d" % (total - success))
    log("  有片区标签: %d" % correct_labeled)
    if correct_labeled > 0:
        acc = correct_matched / correct_labeled * 100
        log("  分类准确率: %.1f%% (%d/%d)" % (acc, correct_matched, correct_labeled))
    log("=" * 70)

    # 按策略统计
    strategy_stats = {}
    for r in results:
        s = r["strategy"]
        if s not in strategy_stats:
            strategy_stats[s] = {"total": 0, "correct": 0, "wrong": 0}
        strategy_stats[s]["total"] += 1
        if r["correct"]:
            strategy_stats[s]["correct"] += 1
        elif r["expected_zone"]:
            strategy_stats[s]["wrong"] += 1

    log("\n  策略使用统计:")
    for s, stats in sorted(strategy_stats.items(), key=lambda x: -x[1]["total"]):
        log("    %s: %d次 (正确%d, 错误%d)" % (s, stats["total"], stats["correct"], stats["wrong"]))

    # 错误详情
    errors = [r for r in results if r["expected_zone"] and not r["correct"]]
    if errors:
        log("\n  ❌ 分类错误详情:")
        for e in errors[:10]:
            log("    %s → %s (期望:%s) 策略:%s" % (
                e["name"][:20] if e["name"] else e["addr"][:20],
                e["matched_zone"], e["expected_zone"], e["strategy"]))

    # 保存结果
    out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "geocode_test_result.txt")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(out_lines))
    log("\n完整报告已保存到: %s" % out_path)

    # 保存JSON格式方便分析
    json_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "geocode_test_result.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    log("JSON数据已保存到: %s" % json_path)

if __name__ == "__main__":
    main()
