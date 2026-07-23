#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
片区分类开发工具 v1.2
改进：GCJ-02坐标 + 高德瓦片，消除转换误差
"""
import math, json, os, time, re, uuid, threading, shutil, tempfile
from datetime import datetime, timezone
from io import BytesIO
import requests
from flask import Flask, request, jsonify, render_template, send_file
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)
app.config["TEMPLATES_AUTO_RELOAD"] = True

@app.after_request
def add_cors_headers(response):
    origin = request.headers.get("Origin", "")
    if "*" in ALLOWED_ORIGINS:
        response.headers["Access-Control-Allow-Origin"] = "*"
    elif origin in ALLOWED_ORIGINS:
        response.headers["Access-Control-Allow-Origin"] = origin
        response.headers["Vary"] = "Origin"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization"
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
    return response

@app.before_request
def handle_cors_preflight():
    if request.method == "OPTIONS":
        return ("", 204)
# 高德 Key 只从服务端环境变量读取，避免密钥进入前端或代码仓库。
GAODE_KEY = os.environ.get("GAODE_KEY", "").strip()
MAX_ROWS = int(os.environ.get("MAX_ROWS", "5000"))
MAX_UPLOAD_BYTES = int(os.environ.get("MAX_UPLOAD_BYTES", str(20 * 1024 * 1024)))
JOB_TTL_SECONDS = int(os.environ.get("JOB_TTL_SECONDS", str(60 * 60)))
AUTO_ACCEPT_MIN = float(os.environ.get("AUTO_ACCEPT_MIN", "0.90"))
BOUNDARY_REVIEW_METERS = float(os.environ.get("BOUNDARY_REVIEW_METERS", "25"))
ALLOWED_ORIGINS = [origin.strip() for origin in os.environ.get("ALLOWED_ORIGINS", "*").split(",") if origin.strip()]
ZONES_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "zones.json")
GEOCODE_CACHE = {}
GAODE_ERROR = threading.local()
JOBS = {}
JOBS_LOCK = threading.RLock()
JOB_ROOT = os.path.join(tempfile.gettempdir(), "shangcheng_zone_jobs")
os.makedirs(JOB_ROOT, exist_ok=True)

def utc_now():
    return datetime.now(timezone.utc).isoformat()

def cleanup_jobs():
    cutoff = time.time() - JOB_TTL_SECONDS
    with JOBS_LOCK:
        stale = [jid for jid, job in JOBS.items() if job.get("updated_ts", 0) < cutoff]
        for jid in stale:
            job = JOBS.pop(jid, None)
            if job:
                shutil.rmtree(job.get("dir", ""), ignore_errors=True)

def get_job(job_id):
    cleanup_jobs()
    with JOBS_LOCK:
        return JOBS.get(job_id)

def create_job(headers, rows, filename):
    job_id = uuid.uuid4().hex
    job_dir = os.path.join(JOB_ROOT, job_id)
    os.makedirs(job_dir, exist_ok=True)
    job = {
        "id": job_id, "dir": job_dir, "filename": filename or "上传文件.xlsx",
        "headers": headers, "rows": rows, "results": [], "zones": [],
        "status": "uploaded", "progress": 0, "message": "等待开始",
        "created_at": utc_now(), "updated_at": utc_now(), "updated_ts": time.time(),
        "name_col": "", "addr_col": "",
    }
    with JOBS_LOCK:
        JOBS[job_id] = job
    return job

def touch_job(job):
    job["updated_at"] = utc_now()
    job["updated_ts"] = time.time()

def set_job(job, **values):
    with JOBS_LOCK:
        job.update(values)
        touch_job(job)

def safe_float(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return None

GAODE_ERROR_MESSAGES = {
    "10001": "Key无效", "10002": "Key无效", "10003": "调用额度不足",
    "10004": "调用过于频繁", "10005": "IP白名单或调用权限错误", "10007": "Key无效",
    "10008": "Key未激活", "10009": "调用权限不足", "10010": "调用额度不足",
}

def set_gaode_error(status=None, info=""):
    GAODE_ERROR.value = {"status": status, "info": info} if status or info else None

def get_gaode_error():
    return getattr(GAODE_ERROR, "value", None)

def classify_gaode_error(data=None, exc=None):
    if exc is not None:
        return {"code": "network", "message": "网络错误", "detail": str(exc)}
    data = data or {}
    code = str(data.get("infocode", ""))
    info = data.get("info", "") or ""
    if code in GAODE_ERROR_MESSAGES:
        return {"code": code, "message": GAODE_ERROR_MESSAGES[code], "detail": info}
    if data.get("status") != "1":
        return {"code": code or "api", "message": "高德接口错误", "detail": info or "未返回有效结果"}
    return {"code": "not_found", "message": "未找到地址", "detail": info}

def status_from_gaode_error(error):
    if not error:
        return "未找到地址"
    if error.get("status") == "network":
        return "网络错误"
    if error.get("status") in GAODE_ERROR_MESSAGES:
        return GAODE_ERROR_MESSAGES[error["status"]]
    return error.get("message", "高德接口错误")
def transform_lat(x, y):
    ret = -100.0 + 2.0*x + 3.0*y + 0.2*y*y + 0.1*x*y + 0.2*abs(x)**0.5
    ret += (20.0*math.sin(6.0*x*math.pi) + 20.0*math.sin(2.0*x*math.pi)) * 2.0/3.0
    ret += (20.0*math.sin(y*math.pi) + 40.0*math.sin(y/3.0*math.pi)) * 2.0/3.0
    ret += (160.0*math.sin(y/12.0*math.pi) + 320.0*math.sin(y*math.pi/30.0)) * 2.0/3.0
    return ret

def transform_lng(x, y):
    ret = 300.0 + x + 2.0*y + 0.1*x*x + 0.1*x*y + 0.1*abs(x)**0.5
    ret += (20.0*math.sin(6.0*x*math.pi) + 20.0*math.sin(2.0*x*math.pi)) * 2.0/3.0
    ret += (20.0*math.sin(x*math.pi) + 40.0*math.sin(x/3.0*math.pi)) * 2.0/3.0
    ret += (150.0*math.sin(x/12.0*math.pi) + 300.0*math.sin(x/30.0*math.pi)) * 2.0/3.0
    return ret

def delta(lng, lat):
    dlat = transform_lat(lng - 105.0, lat - 35.0)
    dlng = transform_lng(lng - 105.0, lat - 35.0)
    radlat = lat / 180.0 * math.pi
    magic = math.sin(radlat)
    magic = 1 - 0.00669342162296594323 * magic * magic
    sqrtmagic = math.sqrt(magic)
    dlat = (dlat * 180.0) / ((6378245.0 * (1 - 0.00669342162296594323)) / (magic * sqrtmagic) * math.pi)
    dlng = (dlng * 180.0) / (6378245.0 / sqrtmagic * math.cos(radlat) * math.pi)
    return dlng, dlat

def gcj02_to_wgs84(lng, lat):
    dlng, dlat = delta(lng, lat)
    return lng - dlng, lat - dlat


def wgs84_to_gcj02(lng, lat):
    """Convert WGS-84 to GCJ-02 (Mars coordinate)"""
    dlng, dlat = delta(lng, lat)
    return lng + dlng, lat + dlat

def point_in_polygon(lng, lat, polygon):
    n = len(polygon)
    inside = False
    j = n - 1
    for i in range(n):
        xi, yi = polygon[i]
        xj, yj = polygon[j]
        if ((yi > lat) != (yj > lat)):
            xinters = (xj - xi) * (lat - yi) / (yj - yi) + xi
            if lng < xinters:
                inside = not inside
        j = i
    return inside

def polygon_area(polygon):
    return abs(sum(polygon[i][0] * polygon[(i + 1) % len(polygon)][1] - polygon[(i + 1) % len(polygon)][0] * polygon[i][1] for i in range(len(polygon))) / 2)

def orientation(a, b, c):
    value = (b[1] - a[1]) * (c[0] - b[0]) - (b[0] - a[0]) * (c[1] - b[1])
    if abs(value) < 1e-12:
        return 0
    return 1 if value > 0 else 2

def segments_intersect(a, b, c, d):
    o1, o2, o3, o4 = orientation(a, b, c), orientation(a, b, d), orientation(c, d, a), orientation(c, d, b)
    if o1 != o2 and o3 != o4:
        return True
    return False

def polygon_self_intersects(polygon):
    n = len(polygon)
    for i in range(n):
        a, b = polygon[i], polygon[(i + 1) % n]
        for j in range(i + 1, n):
            if j in (i, (i + 1) % n, (i - 1) % n):
                continue
            c, d = polygon[j], polygon[(j + 1) % n]
            if segments_intersect(a, b, c, d):
                return True
    return False

def validate_polygon(polygon):
    if not isinstance(polygon, list) or len(polygon) < 3:
        return False, "片区至少需要3个顶点"
    points = []
    for point in polygon:
        if not isinstance(point, (list, tuple)) or len(point) < 2:
            return False, "片区坐标格式不正确"
        lng, lat = safe_float(point[0]), safe_float(point[1])
        if lng is None or lat is None or not (-180 <= lng <= 180 and -90 <= lat <= 90):
            return False, "片区坐标超出范围"
        points.append([lng, lat])
    if polygon_area(points) < 1e-8:
        return False, "片区面积不能为零"
    if polygon_self_intersects(points):
        return False, "片区边界不能自相交"
    return True, points

def zones_overlap(first, second):
    a, b = first.get("polygon", []), second.get("polygon", [])
    if not a or not b:
        return False
    aminx, aminy = min(p[0] for p in a), min(p[1] for p in a)
    amaxx, amaxy = max(p[0] for p in a), max(p[1] for p in a)
    bminx, bminy = min(p[0] for p in b), min(p[1] for p in b)
    bmaxx, bmaxy = max(p[0] for p in b), max(p[1] for p in b)
    if amaxx < bminx or bmaxx < aminx or amaxy < bminy or bmaxy < aminy:
        return False
    for i in range(len(a)):
        for j in range(len(b)):
            if segments_intersect(a[i], a[(i + 1) % len(a)], b[j], b[(j + 1) % len(b)]):
                return True
    return point_in_polygon(a[0][0], a[0][1], b) or point_in_polygon(b[0][0], b[0][1], a)

def point_segment_distance_meters(point, start, end):
    lat_scale = 111320.0
    lng_scale = 111320.0 * math.cos(math.radians(point[1]))
    px, py = point[0] * lng_scale, point[1] * lat_scale
    ax, ay = start[0] * lng_scale, start[1] * lat_scale
    bx, by = end[0] * lng_scale, end[1] * lat_scale
    dx, dy = bx - ax, by - ay
    if dx == 0 and dy == 0:
        return math.hypot(px - ax, py - ay)
    t = max(0, min(1, ((px - ax) * dx + (py - ay) * dy) / (dx * dx + dy * dy)))
    return math.hypot(px - (ax + t * dx), py - (ay + t * dy))

def boundary_distance_meters(lng, lat, polygon):
    point = [lng, lat]
    return min(point_segment_distance_meters(point, polygon[i], polygon[(i + 1) % len(polygon)]) for i in range(len(polygon)))

def geocode_confidence(result):
    if not result:
        return 0.0
    score = 0.55
    level = result.get("level", "")
    strategy = result.get("strategy", "")
    if level in {"门牌号", "兴趣点", "楼栋", "住宅区", "商务写字楼", "公司企业"}:
        score += 0.20
    if strategy.startswith("poi-company"):
        score += 0.12
    if strategy in {"geo-primary", "poi-address"}:
        score += 0.08
    if result.get("regeo_address"):
        score += 0.05
    if result.get("accuracy_warning"):
        score -= 0.25
    return round(max(0.0, min(0.99, score)), 2)
def _clean_address(a):
    """Remove problematic content from address string"""
    a = a.strip()
    a = re.sub(r"\s+", "", a)
    a = re.sub(r"[（）；：【】、，．！？《》、]", "", a)
    return a.strip()

def _poi_search(keywords, city="杭州市", orig_addr=None):
    if not keywords or len(keywords) < 2:
        return None
    try:
        params = {"key": GAODE_KEY, "keywords": keywords, "city": city, "offset": 5, "page": 1, "extensions": "base", "output": "JSON"}
        resp = requests.get("https://restapi.amap.com/v3/place/text", params=params, timeout=10)
        data = resp.json()
        if data.get("status") == "1" and data.get("count"):
            set_gaode_error(None, "")
            count = int(data.get("count", 0))
            if count > 0:
                pois = data.get("pois", [])
                if pois:
                    best = None
                    best_score = -1
                    for poi in pois:
                        loc = poi.get("location", "")
                        if not loc:
                            continue
                        glng, glat = loc.split(",")
                        lng, lat = float(glng), float(glat)
                        entry = {"lng": round(lng, 6), "lat": round(lat, 6), "level": poi.get("typecode", ""), "formatted_address": poi.get("address", ""), "poi_name": poi.get("name", ""), "source": "poi"}
                        typecode = poi.get("typecode", "")
                        score = 0
                        for pref in ["05","06","07"]:
                            if typecode.startswith(pref): score += 3
                        for pref in ["14","15"]:
                            if typecode.startswith(pref): score += 2
                        for pref in ["01","02"]:
                            if typecode.startswith(pref): score -= 2
                        for pref in ["10","11"]:
                            if typecode.startswith(pref): score -= 1
                        poi_addr = poi.get("address", "") or ""
                        name = poi.get("name", "") or ""
                        kw = keywords.strip()
                        # Clean names for comparison (handle parentheses)
                        kw_clean = _clean_address(kw)
                        name_clean = _clean_address(name)
                        # Name matching: prefer exact matches on cleaned names
                        if kw_clean == name_clean:
                            score += 12
                        elif name_clean.startswith(kw_clean) or kw_clean.startswith(name_clean):
                            score += 8
                        elif kw_clean in name_clean:
                            score += 5
                        # Raw substring match
                        if kw in name:
                            score += 3
                        # Address matching
                        if kw in poi_addr:
                            score += 3
                        # Cross-reference with original address if provided
                        if orig_addr:
                            oa = _clean_address(orig_addr)
                            pa = _clean_address(poi_addr)
                            common = set(w for w in oa if len(w) >= 2) & set(w for w in pa if len(w) >= 2)
                            score += len(common) * 0.5
                            m = re.search(r'[路街巷弄]d+[号]', oa)
                            if m and m.group() in poi_addr:
                                score += 5
                        # Prefer POIs with detailed address
                        if len(poi_addr) > 8:
                            score += 1
                        if not best or score > best_score:
                            best = entry
                            best_score = score
                    return best
        set_gaode_error(data.get("infocode"), data.get("info", ""))
        return None
    except Exception as exc:
        set_gaode_error("network", str(exc))
        return None

def _inputtips(keywords, city="杭州市"):
    if not keywords or len(keywords) < 2:
        return None
    try:
        params = {"key": GAODE_KEY, "keywords": keywords, "city": city, "citylimit": True, "output": "JSON"}
        time.sleep(0.1)
        resp = requests.get("https://restapi.amap.com/v3/assistant/inputtips", params=params, timeout=10)
        data = resp.json()
        if data.get("status") == "1" and data.get("tips"):
            set_gaode_error(None, "")
            tips = data["tips"]
            for tip in tips:
                loc = tip.get("location", "")
                if loc:
                    glng, glat = loc.split(",")
                    lng, lat = float(glng), float(glat)
                    return {"lng": round(lng, 6), "lat": round(lat, 6), "level": tip.get("typecode", ""), "formatted_address": tip.get("address", ""), "poi_name": tip.get("name", ""), "source": "tips"}
        set_gaode_error(data.get("infocode"), data.get("info", ""))
        return None
    except Exception as exc:
        set_gaode_error("network", str(exc))
        return None

def _geo(addr, city=""):
    if not addr or len(addr) < 4:
        return None
    try:
        params = {"key": GAODE_KEY, "address": addr, "output": "JSON"}
        if city:
            params["city"] = city
        resp = requests.get("https://restapi.amap.com/v3/geocode/geo", params=params, timeout=10)
        data = resp.json()
        if data.get("status") == "1" and data.get("geocodes"):
            set_gaode_error(None, "")
            for gc in data["geocodes"]:
                loc = gc.get("location", "")
                if loc:
                    glng, glat = loc.split(",")
                    lng, lat = float(glng), float(glat)
                    return {"lng": round(lng, 6), "lat": round(lat, 6), "level": gc.get("level", ""), "formatted_address": gc.get("formatted_address", addr), "source": "geo"}
        set_gaode_error(data.get("infocode"), data.get("info", ""))
        return None
    except Exception as exc:
        set_gaode_error("network", str(exc))
        return None

def _regeo(lng, lat):
    """Reverse geocode - get formatted address from coordinates"""
    try:
        params = {"key": GAODE_KEY, "location": f"{lng},{lat}", "radius": 100, "extensions": "base", "output": "JSON"}
        resp = requests.get("https://restapi.amap.com/v3/geocode/regeo", params=params, timeout=10)
        data = resp.json()
        if data.get("status") == "1":
            set_gaode_error(None, "")
            formatted = data.get("regeocode", {}).get("formatted_address", "")
            if formatted:
                return formatted
        set_gaode_error(data.get("infocode"), data.get("info", ""))
        return ""
    except Exception as exc:
        set_gaode_error("network", str(exc))
        return ""

def geocode_address(address, company_name=""):
    orig_addr = address.strip()
    if not orig_addr or len(orig_addr) < 4:
        return None
    cache_key = company_name + "|" + orig_addr
    if cache_key in GEOCODE_CACHE:
        return GEOCODE_CACHE[cache_key]
    set_gaode_error(None, "")
    cleaned = orig_addr
    try:
        cleaned = _clean_address(orig_addr)
    except:
        pass
    fallback = None

    def _district_of(addr):
        for d in ["上城区", "下城区", "西湖区", "拱墅区", "滨江区", "萧山区", "余杭区", "临平区", "钱塘区", "富阳区", "临安区", "桐庐县", "淳安县", "建德市"]:
            if d in addr:
                return d
        return None

    def _addr_ok(result):
        if not result: return None
        fa = result.get("formatted_address", "") or ""
        orig_district = _district_of(orig_addr) or _district_of(cleaned)
        if not orig_district: return None
        if _district_of(fa) == orig_district: return True
        pn = result.get("poi_name", "") or ""
        if _district_of(pn) == orig_district: return True
        # If both result fields lack district info, we can't determine
        # if it's wrong - allow it through rather than rejecting
        if not _district_of(fa) and not _district_of(pn):
            return None
        return False

    def _try(name, fn):
        nonlocal fallback
        try:
            result = fn()
            if result and result.get("lng"):
                result["strategy"] = name
                ok = _addr_ok(result)
                result["accuracy_warning"] = (ok is False)
                if ok is not False:
                    return result
                if not fallback:
                    fallback = dict(result)
                    fallback["accuracy_warning"] = True
                    fallback["strategy"] = name
        except:
            pass
        return None

    if cleaned and len(cleaned) >= 6:
        r = _try("geo-primary", lambda: _geo(cleaned, "杭州"))
        if r:
            # Reject coarse geo results (road/district level) so
            # company name POI search can be tried instead
            geo_level = r.get("level", "")
            coarse_levels = {"区县", "城市", "道路", "乡镇", "村庄"}
            if geo_level in coarse_levels and company_name and len(company_name.strip()) >= 4:
                if not fallback:
                    fallback = dict(r)
                    fallback["accuracy_warning"] = True
                    fallback["strategy"] = "geo-primary(coarse)"
                r = None
        if r:
            try:
                rg_addr = _regeo(r["lng"], r["lat"])
                if rg_addr:
                    r["regeo_address"] = rg_addr
            except:
                pass
            GEOCODE_CACHE[cache_key] = r
            return r

    # Search company name as POI first - often more accurate than address geocoding
    if company_name and len(company_name.strip()) >= 4:
        try:
            cn = _clean_address(company_name)
            if len(cn) >= 4:
                r = _try("poi-company-first", lambda: _poi_search(cn, "杭州市", orig_addr))
                if r:
                    try:
                        rg_addr = _regeo(r["lng"], r["lat"])
                        if rg_addr:
                            r["regeo_address"] = rg_addr
                    except:
                        pass
                    GEOCODE_CACHE[cache_key] = r
                    return r
        except:
            pass

    # Search company name with district context for higher accuracy
    if company_name and len(company_name.strip()) >= 4:
        try:
            cn = _clean_address(company_name)
            district = None
            for d in ["上城区", "下城区", "西湖区", "拱墅区", "滨江区", "萧山区", "余杭区", "临平区", "钱塘区", "富阳区", "临安区", "桐庐县", "淳安县", "建德市"]:
                if d in orig_addr:
                    district = d
                    break
            if district and len(cn) >= 4:
                district_kw = district + cn
                r = _try("poi-company-district", lambda: _poi_search(district_kw, "杭州市", orig_addr))
                if r:
                    try:
                        rg_addr = _regeo(r["lng"], r["lat"])
                        if rg_addr:
                            r["regeo_address"] = rg_addr
                    except:
                        pass
                    GEOCODE_CACHE[cache_key] = r
                    return r
        except:
            pass

    st = re.search(r"[路街巷弄]\d+[号]", cleaned) if cleaned else None
    if st:
        sk = st.group()
        for d in ["上城区", "下城区", "西湖区", "拱墅区", "滨江区"]:
            if d in cleaned:
                sk = d + sk
                break
        r = _try("poi-streetnum", lambda: _poi_search(sk, "杭州市", orig_addr))
        if r:
            try:
                rg_addr = _regeo(r["lng"], r["lat"])
                if rg_addr:
                    r["regeo_address"] = rg_addr
            except:
                pass
            GEOCODE_CACHE[cache_key] = r
            return r

    if cleaned and len(cleaned) >= 6:
        r = _try("tips-address", lambda: _inputtips(cleaned[:40], "杭州市"))
        if r:
            try:
                rg_addr = _regeo(r["lng"], r["lat"])
                if rg_addr:
                    r["regeo_address"] = rg_addr
            except:
                pass
            GEOCODE_CACHE[cache_key] = r
            return r

    r = _try("poi-address", lambda: _poi_search(cleaned, "杭州市", orig_addr))
    if r:
        try:
            rg_addr = _regeo(r["lng"], r["lat"])
            if rg_addr:
                r["regeo_address"] = rg_addr
        except:
            pass
        GEOCODE_CACHE[cache_key] = r
        return r

    if company_name and len(company_name) >= 4:
        r = _try("poi-company", lambda: _poi_search(company_name.strip(), "杭州市", orig_addr))
        if r:
            try:
                rg_addr = _regeo(r["lng"], r["lat"])
                if rg_addr:
                    r["regeo_address"] = rg_addr
            except:
                pass
            GEOCODE_CACHE[cache_key] = r
            return r

    if company_name and len(company_name) >= 2:
        combined = company_name.strip() + " " + cleaned[:30]
        r = _try("poi-combined", lambda: _poi_search(combined, "杭州市", orig_addr))
        if r:
            try:
                rg_addr = _regeo(r["lng"], r["lat"])
                if rg_addr:
                    r["regeo_address"] = rg_addr
            except:
                pass
            GEOCODE_CACHE[cache_key] = r
            return r

    # Try geo with uncleaned address (retains more context) and specific city
    r = _try("geo-origaddr", lambda: _geo(orig_addr[:60], "杭州市"))
    if r:
        try:
            rg_addr = _regeo(r["lng"], r["lat"])
            if rg_addr:
                r["regeo_address"] = rg_addr
        except:
            pass
        GEOCODE_CACHE[cache_key] = r
        return r

    # Try POI search with uncleaned full address (may contain useful details)
    r = _try("poi-address-full", lambda: _poi_search(orig_addr[:60], "杭州市", orig_addr))
    if r:
        try:
            rg_addr = _regeo(r["lng"], r["lat"])
            if rg_addr:
                r["regeo_address"] = rg_addr
        except:
            pass
        GEOCODE_CACHE[cache_key] = r
        return r

    if fallback and fallback.get("lng"):
        try:
            rg_addr = _regeo(fallback["lng"], fallback["lat"])
            if rg_addr:
                fallback["regeo_address"] = rg_addr
        except:
            pass
        GEOCODE_CACHE[cache_key] = fallback
        return fallback

    for av in [cleaned, orig_addr]:
        try:
            cs = av[:50]
            if len(cs) >= 6:
                rg = _geo(cs, "杭州")
                if not rg or not rg.get("lng"):
                    rg = _geo(cs, "")
                if rg and rg.get("lng"):
                    rg["strategy"] = "geo-fallback"
                    rg["accuracy_warning"] = True
                    GEOCODE_CACHE[cache_key] = rg
                    return rg
        except:
            continue

    GEOCODE_CACHE[cache_key] = None
    return None

def load_zones():
    if os.path.exists(ZONES_FILE):
        with open(ZONES_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return []

def save_zones(zones):
    with open(ZONES_FILE, "w", encoding="utf-8") as f:
        json.dump(zones, f, ensure_ascii=False, indent=2)

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/healthz")
def healthz():
    return jsonify({"status": "ok", "gaode_configured": bool(GAODE_KEY),
                    "gaode_message": "已配置" if GAODE_KEY else "Key未配置",
                    "max_rows": MAX_ROWS})

@app.route("/api/upload", methods=["POST"])
def upload_excel():
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "未上传文件"}), 400
    if not (file.filename or "").lower().endswith(".xlsx"):
        return jsonify({"error": "仅支持 .xlsx 文件"}), 400
    file.stream.seek(0, os.SEEK_END)
    if file.stream.tell() > MAX_UPLOAD_BYTES:
        return jsonify({"error": f"文件不能超过 {MAX_UPLOAD_BYTES // 1024 // 1024}MB"}), 413
    file.stream.seek(0)
    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    except Exception as exc:
        return jsonify({"error": f"Excel读取失败：{exc}"}), 400
    ws = wb.active
    if ws.max_row - 1 > MAX_ROWS:
        return jsonify({"error": f"单个文件最多支持 {MAX_ROWS} 行数据"}), 413
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        headers.append(str(v) if v else "列" + str(c))
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            row[headers[c - 1]] = str(v) if v is not None else ""
        rows.append(row)
    job = create_job(headers, rows, file.filename)
    return jsonify({"job_id": job["id"], "headers": headers, "rows": rows, "total": len(rows)})

def _job_summary(job):
    results = job.get("results", [])
    return {
        "total": len(results),
        "automatic": sum(1 for row in results if row.get("review_status") == "自动通过"),
        "review": sum(1 for row in results if row.get("review_status") == "待复核"),
        "failed": sum(1 for row in results if row.get("review_status") == "无法定位"),
        "outside": sum(1 for row in results if row.get("review_status") == "片区外" or row.get("zone") == "片区外"),
        "manual": sum(1 for row in results if row.get("manual_reviewed")),
    }

def _run_geocode_job(job, name_col, addr_col):
    set_job(job, status="geocoding", progress=0, message="正在进行地址解析", name_col=name_col, addr_col=addr_col)
    rows = job.get("rows", [])
    results = []
    for index, item in enumerate(rows):
        address = str(item.get(addr_col, "") or "").strip()
        company_name = str(item.get(name_col, "") or "").strip() if name_col else ""
        base = {**item, "row_index": index}
        if not address:
            base.update({"lng": None, "lat": None, "geocode_status": "地址为空", "confidence": 0, "review_status": "无法定位", "source": "", "error_detail": "地址列为空"})
        elif not GAODE_KEY:
            base.update({"lng": None, "lat": None, "geocode_status": "Key未配置", "confidence": 0, "review_status": "无法定位", "source": "", "error_detail": "请通过统一启动脚本启动服务"})
        else:
            result = geocode_address(address, company_name)
            if result and result.get("lng") is not None:
                confidence = geocode_confidence(result)
                base.update({"lng": result["lng"], "lat": result["lat"], "geocode_status": "成功",
                             "level": result.get("level", ""), "formatted_address": result.get("formatted_address", address),
                             "source": result.get("source", ""), "strategy": result.get("strategy", ""),
                             "poi_name": result.get("poi_name", ""), "accuracy_warning": result.get("accuracy_warning", False),
                             "regeo_address": result.get("regeo_address", ""), "confidence": confidence,
                             "review_status": "待复核" if confidence < AUTO_ACCEPT_MIN else "自动通过", "manual_reviewed": False})
            else:
                error = get_gaode_error()
                status = status_from_gaode_error(error)
                base.update({"lng": None, "lat": None, "geocode_status": status, "confidence": 0,
                             "review_status": "无法定位", "source": "", "error_detail": (error or {}).get("detail", "")})
        results.append(base)
        if index % 25 == 0 or index == len(rows) - 1:
            set_job(job, results=results, progress=int((index + 1) / max(1, len(rows)) * 100), message=f"已处理 {index + 1}/{len(rows)} 行")
        if index > 0 and GAODE_KEY:
            time.sleep(0.3)
    set_job(job, results=results, status="geocoded", progress=100, message="地址解析完成")

@app.route("/api/tasks/<job_id>/geocode", methods=["POST"])
def start_geocode_task(job_id):
    job = get_job(job_id)
    if not job:
        return jsonify({"error": "任务不存在或已过期"}), 404
    data = request.get_json(silent=True) or {}
    name_col, addr_col = data.get("name_col", ""), data.get("addr_col", "")
    if not addr_col or addr_col not in job["headers"]:
        return jsonify({"error": "请选择有效的地址列"}), 400
    if job.get("status") == "geocoding":
        return jsonify({"status": "geocoding"}), 202
    threading.Thread(target=_run_geocode_job, args=(job, name_col, addr_col), daemon=True).start()
    return jsonify({"status": "geocoding", "job_id": job_id}), 202

@app.route("/api/tasks/<job_id>", methods=["GET"])
def task_status(job_id):
    job = get_job(job_id)
    if not job:
        return jsonify({"error": "任务不存在或已过期"}), 404
    return jsonify({"job_id": job_id, "status": job["status"], "progress": job["progress"],
                    "message": job["message"], "headers": job["headers"], "results": job.get("results", []),
                    "summary": _job_summary(job), "zones": job.get("zones", [])})

@app.route("/api/geocode", methods=["POST"])
def batch_geocode():
    data = request.get_json()
    items = data.get("items", [])
    addr_col = data.get("addr_col", "")
    name_col = data.get("name_col", "")
    results = []
    total = len(items)
    for i, item in enumerate(items):
        address = item.get(addr_col, "")
        company_name = item.get(name_col, "") if name_col else ""
        if not address:
            results.append({**item, "lng": None, "lat": None, "geocode_status": "地址为空"})
            continue
        result = geocode_address(address, company_name)
        if result and result.get("lng"):
            results.append({**item,
                "lng": result["lng"], "lat": result["lat"],
                "geocode_status": "成功",
                "level": result.get("level", ""),
                "formatted_address": result.get("formatted_address", address),
                "source": result.get("source", ""),
                "strategy": result.get("strategy", ""),
                "poi_name": result.get("poi_name", ""),
                "accuracy_warning": result.get("accuracy_warning", False),
                "regeo_address": result.get("regeo_address", "")})
        else:
            error = get_gaode_error()
            results.append({**item, "lng": None, "lat": None,
                "geocode_status": status_from_gaode_error(error), "error_detail": (error or {}).get("detail", ""),
                "level": "", "formatted_address": "",
                "source": "", "strategy": "", "poi_name": "",
                "accuracy_warning": False,
                "regeo_address": ""})
        if i > 0:
            time.sleep(0.3)
    sc = sum(1 for r in results if r.get("lng"))
    return jsonify({"results": results, "total": total, "success": sc, "failed": total - sc})

@app.route("/api/classify", methods=["POST"])
def classify_points():
    data = request.get_json()
    points = data.get("points", [])
    zones = data.get("zones", [])
    results = []
    overlap_names = []
    for i, first in enumerate(zones):
        for second in zones[i + 1:]:
            if zones_overlap(first, second):
                overlap_names.extend([first.get("name", ""), second.get("name", "")])
    for pt in points:
        lng, lat = pt.get("lng"), pt.get("lat")
        zn = "未划分"
        zc = None
        if lng is not None and lat is not None:
            for z in zones:
                poly = z.get("polygon", [])
                if len(poly) >= 3 and point_in_polygon(lng, lat, poly):
                    zn = z.get("name", "未命名片区")
                    zc = z.get("color", "#3388ff")
                    break
        if zn == "未划分" and lng is not None and lat is not None:
            zn = "片区外"
        distance = boundary_distance_meters(lng, lat, next((z.get("polygon", []) for z in zones if z.get("name") == zn), [])) if zn not in {"未划分", "片区外"} else None
        status = pt.get("review_status", "待复核")
        if zn == "未划分":
            status = "待复核" if lng is not None else "无法定位"
        elif zn == "片区外":
            status = "片区外"
        elif distance is not None and distance <= BOUNDARY_REVIEW_METERS:
            status = "边界待确认"
        elif pt.get("confidence", 0) >= AUTO_ACCEPT_MIN and not pt.get("accuracy_warning") and not overlap_names:
            status = "自动通过"
        else:
            status = "待复核"
        results.append({**pt, "zone": zn, "zone_color": zc, "boundary_distance_m": round(distance, 1) if distance is not None else None, "review_status": status, "overlap_warning": bool(overlap_names)})
    return jsonify({"results": results, "overlap_zones": sorted(set(overlap_names))})

@app.route("/api/tasks/<job_id>/classify", methods=["POST"])
def classify_task(job_id):
    job = get_job(job_id)
    if not job:
        return jsonify({"error": "任务不存在或已过期"}), 404
    data = request.get_json(silent=True) or {}
    zones, errors = normalize_zones(data.get("zones", []))
    if errors:
        return jsonify({"error": "；".join(errors)}), 400
    if not job.get("results"):
        return jsonify({"error": "请先完成地址解析"}), 400
    classified = classify_points_data(job["results"], zones)
    set_job(job, results=classified, zones=zones, status="classified", progress=100, message="分类完成")
    return jsonify({"results": classified, "summary": _job_summary(job)})

def normalize_zones(raw_zones):
    zones, errors = [], []
    if not isinstance(raw_zones, list):
        return [], ["片区格式不正确"]
    names = set()
    for raw in raw_zones:
        name = str(raw.get("name", "")).strip()
        if not name or name in names:
            errors.append("片区名称不能为空且不能重复")
            continue
        ok, polygon = validate_polygon(raw.get("polygon", []))
        if not ok:
            errors.append(f"{name}：{polygon}")
            continue
        zone = {"name": name, "color": raw.get("color", "#3388ff"), "polygon": polygon}
        if any(zones_overlap(zone, other) for other in zones):
            errors.append(f"{name}：与已有片区重叠")
            continue
        zones.append(zone)
        names.add(name)
    return zones, errors

def classify_points_data(points, zones):
    output = []
    for pt in points:
        lng, lat = safe_float(pt.get("lng")), safe_float(pt.get("lat"))
        zone, color, distance = "未划分", None, None
        for z in zones:
            if lng is not None and lat is not None and point_in_polygon(lng, lat, z["polygon"]):
                zone, color = z["name"], z["color"]
                distance = boundary_distance_meters(lng, lat, z["polygon"])
                break
        if lng is None or lat is None:
            review_status = "无法定位"
        elif zone == "未划分":
            zone = "片区外"
            review_status = "片区外"
        elif distance <= BOUNDARY_REVIEW_METERS or pt.get("accuracy_warning") or pt.get("confidence", 0) < AUTO_ACCEPT_MIN:
            review_status = "边界待确认" if distance <= BOUNDARY_REVIEW_METERS else "待复核"
        else:
            review_status = "自动通过"
        output.append({**pt, "lng": lng, "lat": lat, "zone": zone, "zone_color": color,
                       "boundary_distance_m": round(distance, 1) if distance is not None else None,
                       "review_status": review_status, "manual_reviewed": bool(pt.get("manual_reviewed", False))})
    return output

@app.route("/api/tasks/<job_id>/review", methods=["POST"])
def review_task(job_id):
    job = get_job(job_id)
    if not job:
        return jsonify({"error": "任务不存在或已过期"}), 404
    data = request.get_json(silent=True) or {}
    updates = data.get("updates", [])
    by_index = {int(row.get("row_index")): row for row in updates if str(row.get("row_index", "")).isdigit()}
    zones = job.get("zones", [])
    zone_map = {z["name"]: z for z in zones}
    for row in job.get("results", []):
        update = by_index.get(row.get("row_index"))
        if not update:
            continue
        lng, lat = safe_float(update.get("lng", row.get("lng"))), safe_float(update.get("lat", row.get("lat")))
        zone_name = str(update.get("zone", row.get("zone", "未划分"))).strip() or "未划分"
        if zone_name not in {"未划分", "片区外"} and zone_name not in zone_map:
            return jsonify({"error": f"不存在的片区：{zone_name}"}), 400
        row.update({"lng": lng, "lat": lat, "zone": zone_name, "zone_color": zone_map.get(zone_name, {}).get("color"),
                    "manual_reviewed": True, "review_status": "人工确认", "review_note": str(update.get("note", ""))[:500]})
    set_job(job, status="reviewed", message="人工复核已保存")
    return jsonify({"results": job["results"], "summary": _job_summary(job)})

@app.route("/api/zones", methods=["GET"])
def get_zones():
    return jsonify({"zones": load_zones()})

@app.route("/api/zones", methods=["POST"])
def save_zones_route():
    raw = (request.get_json(silent=True) or {}).get("zones", [])
    zones, errors = normalize_zones(raw)
    if errors:
        return jsonify({"error": "；".join(errors)}), 400
    payload = {"updated_at": utc_now(), "zones": zones}
    save_zones(zones)
    return jsonify({"status": "ok", "count": len(zones), "updated_at": payload["updated_at"]})

def build_export(results, headers, zone_color_map):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "片区分类结果"
    extra_headers = ["所属片区", "经度", "纬度", "地理编码状态", "置信度", "匹配类型", "匹配来源", "边界距离(米)", "是否人工确认", "复核备注"]
    eh = headers + extra_headers
    hf = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    hfn = Font(color="FFFFFF", bold=True, size=11)
    for ci, hdr in enumerate(eh, 1):
        cell = ws.cell(row=1, column=ci, value=hdr)
        cell.fill, cell.font = hf, hfn
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for ri, item in enumerate(results, 2):
        for ci, hdr in enumerate(headers, 1):
            ws.cell(row=ri, column=ci, value=item.get(hdr, ""))
        values = [item.get("zone", "未划分"), item.get("lng", ""), item.get("lat", ""), item.get("geocode_status", ""),
                  item.get("confidence", ""), item.get("level", ""), item.get("source", ""), item.get("boundary_distance_m", ""),
                  "是" if item.get("manual_reviewed") else "否", item.get("review_note", "")]
        for offset, value in enumerate(values, len(headers) + 1):
            ws.cell(row=ri, column=offset, value=value)
        zone_cell = ws.cell(row=ri, column=len(headers) + 1)
        color = zone_color_map.get(item.get("zone", ""), "F4CCCC").lstrip("#")
        zone_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for cc in range(1, len(eh) + 1):
        values = [ws.cell(row=rr, column=cc).value for rr in range(1, len(results) + 2)]
        ml = max([len(str(v)) for v in values if v is not None] or [8])
        ws.column_dimensions[openpyxl.utils.get_column_letter(cc)].width = min(max(10, ml + 3), 35)
    ws.freeze_panes = "A2"
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

@app.route("/api/tasks/<job_id>/export", methods=["GET"])
def export_task(job_id):
    job = get_job(job_id)
    if not job:
        return jsonify({"error": "任务不存在或已过期"}), 404
    if not job.get("results"):
        return jsonify({"error": "暂无可导出的结果"}), 400
    zone_color_map = {z.get("name"): z.get("color", "#3388ff") for z in job.get("zones", [])}
    output = build_export(job["results"], job["headers"], zone_color_map)
    response = send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="上城区片区分类结果.xlsx")
    response.call_on_close(lambda: cleanup_job(job_id))
    return response

def cleanup_job(job_id):
    with JOBS_LOCK:
        job = JOBS.pop(job_id, None)
    if job:
        shutil.rmtree(job.get("dir", ""), ignore_errors=True)

@app.route("/api/export", methods=["POST"])
def export_results():
    from io import BytesIO
    data = request.get_json()
    results = data.get("results", [])
    headers = data.get("headers", [])
    zcm = data.get("zone_color_map", {})
    output = build_export(results, headers, zcm)
    return send_file(output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name="片区分类结果.xlsx")

if __name__ == "__main__":
    import sys
    port = 5050
    for i, arg in enumerate(sys.argv):
        if arg == "--port" and i + 1 < len(sys.argv):
            try:
                port = int(sys.argv[i + 1])
            except:
                pass
        elif arg.startswith("--port="):
            try:
                port = int(arg.split("=")[1])
            except:
                pass
    port_env = os.environ.get("PORT")
    if port_env:
        try:
            port = int(port_env)
        except:
            pass
    print("=" * 55)
    print("  片区分类开发工具 v1.3")
    print(f"  启动地址: http://localhost:{port}")
    print("  GCJ-02坐标 + 高德地图瓦片")
    print("=" * 55)
    app.run(host="0.0.0.0", port=port, debug=False)
